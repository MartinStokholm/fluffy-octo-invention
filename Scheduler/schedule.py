import sys
import json
import random
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def load_config(config_path):
    try:
        with open(config_path, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Configuration file '{config_path}' not found.")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error parsing '{config_path}': {e}")
        sys.exit(1)

def is_holiday_day(date_obj, holidays):
    return any(date_obj.date() == h.date() for h in holidays)

def can_work_day(person, date_obj):
    return date_obj.strftime('%A') not in person.get('unavailable_days', [])

def not_incompatible(p1, p2):
    return (p1['name'] not in p2.get('incompatible_with', [])) and \
           (p2['name'] not in p1.get('incompatible_with', []))

def add_months(start_date, months):
    month = start_date.month - 1 + months
    year = start_date.year + month // 12
    month = month % 12 + 1
    day = min(start_date.day, [31,
          29 if year%4==0 and not year%100==0 or year%400==0 else 28,
          31,30,31,30,31,31,30,31,30,31][month-1])
    return datetime(year, month, day)

def main():
    # Define the order of days
    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    # Check for correct usage
    if len(sys.argv) < 3:
        print("Usage: python schedule.py YYYY-MM-DD number_of_months")
        sys.exit(1)

    start_date_str = sys.argv[1]
    months_str = sys.argv[2]

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    except ValueError:
        print("Incorrect date format. Please use YYYY-MM-DD.")
        sys.exit(1)

    try:
        number_of_months = int(months_str)
        if number_of_months < 1:
            raise ValueError
    except ValueError:
        print("Number of months must be a positive integer.")
        sys.exit(1)

    # Load configuration
    config = load_config('personas.json')
    personas = config.get('personas', [])
    holidays = [datetime.strptime(h, '%Y-%m-%d') for h in config.get('holidays', [])]

    if not personas:
        print("No personas found in configuration.")
        sys.exit(1)

    # Initialize next_available for each persona
    for p in personas:
        p['next_available'] = datetime.min

    # Define the schedule period based on the number of months
    end_date = add_months(start_date, number_of_months)
    assignments = []
    errors = []

    # Initialize weekend assignment counts with week numbers
    weekend_counts = {persona['name']: {'count': 0, 'weeks': []} for persona in personas}

    current = start_date
    while current <= end_date:
        if not is_holiday_day(current, holidays):
            # Gather available candidates
            candidates = [p for p in personas
                          if p['next_available'] <= current and can_work_day(p, current)]
            
            # Ensure at least two candidates are available
            if len(candidates) < 2:
                errors.append(f"{current.strftime('%Y-%m-%d')} ({current.strftime('%A')}): Not enough available people.")
                assignments.append({
                    'date': current.strftime('%Y-%m-%d'),
                    'person1': "Unassigned",
                    'person2': "Unassigned"
                })
                current += timedelta(days=1)
                continue

            # Attempt to find a compatible pair
            possible_pairs = []
            for i in range(len(candidates)):
                for j in range(i + 1, len(candidates)):
                    if not_incompatible(candidates[i], candidates[j]):
                        possible_pairs.append((candidates[i], candidates[j]))

            if not possible_pairs:
                errors.append(f"{current.strftime('%Y-%m-%d')} ({current.strftime('%A')}): No compatible person pairs available.")
                assignments.append({
                    'date': current.strftime('%Y-%m-%d'),
                    'person1': "Unassigned",
                    'person2': "Unassigned"
                })
                current += timedelta(days=1)
                continue

            # Select a pair
            if current.weekday() < 5:
                # Weekday: Select the first available compatible pair
                chosen_pair = possible_pairs[0]
            else:
                # Weekend: Shuffle and select a random compatible pair
                random.shuffle(possible_pairs)
                chosen_pair = possible_pairs[0]

            assignments.append({
                'date': current.strftime('%Y-%m-%d'),
                'person1': chosen_pair[0]['name'],
                'person2': chosen_pair[1]['name']
            })

            # Update next_available for the assigned people
            chosen_pair[0]['next_available'] = current + timedelta(days=5)
            chosen_pair[1]['next_available'] = current + timedelta(days=5)

            # Update weekend counts if it's a weekend
            if current.weekday() >= 5:  # Saturday=5, Sunday=6
                week_number = current.isocalendar()[1]
                weekend_counts[chosen_pair[0]['name']]['count'] += 1
                weekend_counts[chosen_pair[0]['name']]['weeks'].append(week_number)
                weekend_counts[chosen_pair[1]['name']]['count'] += 1
                weekend_counts[chosen_pair[1]['name']]['weeks'].append(week_number)

        else:
            # Holidays: No assignment needed
            assignments.append({
                'date': current.strftime('%Y-%m-%d'),
                'person1': "Holiday",
                'person2': "Holiday"
            })

        current += timedelta(days=1)

    # Proceed to generate the Excel spreadsheet
    df = pd.DataFrame(assignments)
    df['date_obj'] = pd.to_datetime(df['date'])
    df['week_number'] = df['date_obj'].dt.isocalendar().week
    df['weekday'] = df['date_obj'].dt.day_name()
    df = df.sort_values(by=['week_number', 'date_obj'])

    # Group assignments by week
    weeks_data = {}
    for _, row in df.iterrows():
        w = int(row['week_number'])
        day = row['weekday']  # Monday, Tuesday, ...
        date_str = row['date']  # 'YYYY-MM-DD'
        assigned = f"{row['person1']}\n{row['person2']}"
        if w not in weeks_data:
            weeks_data[w] = {}
        weeks_data[w][day] = {'date': date_str, 'assignment': assigned}
    
    # Ensure all days are present in each week
    for week in weeks_data:
        for day in days_order:
            if day not in weeks_data[week]:
                # Assign default values for missing days
                weeks_data[week][day] = {'date': 'N/A', 'assignment': 'No Assignment'}

    # Create Excel workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Schedule"

    row_idx = 1

    # Define border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Define fill colors
    holiday_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    unassigned_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow

    for week in sorted(weeks_data.keys()):
        # Merge cells for "Week X"
        sheet.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=len(days_order))
        week_cell = sheet.cell(row=row_idx, column=1, value=f"Week {week}")
        
        # Apply bold font and centered alignment to the merged cell
        week_cell.font = Font(bold=True)
        week_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        week_cell.border = thin_border  # Apply border
        
        row_idx += 1

        # Add Date Row
        for col_idx, day_name in enumerate(days_order, start=1):
            date = weeks_data[week][day_name]['date']
            date_cell = sheet.cell(row=row_idx, column=col_idx, value=date)
            date_cell.font = Font(bold=True)
            date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            date_cell.border = thin_border  # Apply border
        row_idx += 1

        # Create weekday header row
        for col_idx, day_name in enumerate(days_order, start=1):
            header_cell = sheet.cell(row=row_idx, column=col_idx, value=day_name)
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_cell.border = thin_border  # Apply border
        row_idx += 1

        # Fill row with assigned pairs
        for col_idx, day_name in enumerate(days_order, start=1):
            assignment = weeks_data[week].get(day_name, {"assignment": "No Assignment"})['assignment']
            assignment_cell = sheet.cell(
                row=row_idx,
                column=col_idx,
                value=assignment
            )
            assignment_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            assignment_cell.border = thin_border  # Apply border

            # Apply fill color based on assignment status
            if assignment == "Holiday":
                assignment_cell.fill = holiday_fill
            elif assignment == "No Assignment":
                assignment_cell.fill = unassigned_fill

        row_idx += 1  # Move to the next row for the next week

    # Add Stats Rows
    row_idx += 1  # Add an empty row before stats
    stats_title = f"Weekend Assignment Stats (Total Weeks: {len(weeks_data)})"
    sheet.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=3)
    stats_cell = sheet.cell(row=row_idx, column=1, value=stats_title)
    stats_cell.font = Font(bold=True)
    stats_cell.alignment = Alignment(horizontal='left', vertical='center')
    row_idx += 1

    # Header for stats
    sheet.cell(row=row_idx, column=1, value="Person Name").font = Font(bold=True)
    sheet.cell(row=row_idx, column=2, value="Weekend Assignments").font = Font(bold=True)
    sheet.cell(row=row_idx, column=3, value="Week Numbers").font = Font(bold=True)
    for col in range(1, 4):
        cell = sheet.cell(row=row_idx, column=col)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
    row_idx += 1

    # Populate stats
    for person, data in weekend_counts.items():
        sheet.cell(row=row_idx, column=1, value=person).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=row_idx, column=2, value=data['count']).alignment = Alignment(horizontal='left', vertical='center')
        week_numbers = ", ".join(map(str, sorted(set(data['weeks']))))
        sheet.cell(row=row_idx, column=3, value=week_numbers).alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 4):
            cell = sheet.cell(row=row_idx, column=col)
            cell.border = thin_border
        row_idx += 1

    # Adjust column widths for better readability
    # Adjust Schedule Sheet Columns
    for i, column in enumerate(days_order, start=1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = 20  # Adjust the width as needed

    # Adjust Stats Section Columns
    sheet.column_dimensions[get_column_letter(1)].width = 20  # Person Name
    sheet.column_dimensions[get_column_letter(2)].width = 25  # Weekend Assignments
    sheet.column_dimensions[get_column_letter(3)].width = 30  # Week Numbers

    # Adjust row heights 
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 30  # Set row height to 30

    # Save as .xlsx
    try:
        wb.save("people_schedule.xlsx")
        print("Schedule successfully generated and saved as 'people_schedule.xlsx'.")
    except Exception as e:
        print(f"Failed to save the schedule: {e}")
        sys.exit(1)

    # Log errors to a separate file
    if errors:
        try:
            with open("schedule_errors.log", "w") as log_file:
                log_file.write("Scheduling completed with some issues:\n")
                log_file.write("The following days could not be fully assigned and need manual handling:\n")
                for error in errors:
                    log_file.write(f" - {error}\n")
            print("\nSome days could not be fully assigned. See 'schedule_errors.log' for details.")
        except Exception as e:
            print(f"Failed to write error log: {e}")
    else:
        print("\nAll days were successfully scheduled.")

if __name__ == "__main__":
    main()