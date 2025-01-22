import json
import os
from datetime import datetime
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, asdict
from utils import ensure_dir_exists
import logging


@dataclass
class PersonResult:
    name: str
    working_day: str
    absence_days: List[str]
    incompatible_with: List[str]
    assigned_shifts: List[str]
    fridays_count: int
    saturdays_count: int
    sundays_count: int

    @property
    def weekday_shifts(self) -> int:
        """
        Calculate the number of assigned shifts excluding weekends.
        """
        return len(self.assigned_shifts) - (
            self.fridays_count + self.saturdays_count + self.sundays_count
        )

    @property
    def weekend_shifts(self) -> int:
        """
        Calculate the number of assigned shifts on weekends.
        """
        return self.fridays_count + self.saturdays_count + self.sundays_count

    @property
    def total_shifts(self) -> int:
        """
        Calculate the total number of assigned shifts.
        """
        return len(self.assigned_shifts)


class ScheduleExporter:
    def __init__(self, json_filepath: str, output_excel: str):
        self.json_filepath = json_filepath
        self.output_excel = output_excel
        self.people: List[PersonResult] = []
        self.schedule: Dict[str, Dict[str, List[str]]] = {}  # {week: {day: [names]}}

    def load_data(self):
        with open(self.json_filepath, "r") as file:
            data = json.load(file)
            for person_data in data:
                person = PersonResult(
                    name=person_data["name"],
                    working_day=person_data["working_day"],
                    absence_days=person_data.get("absence_days", []),
                    incompatible_with=person_data.get("incompatible_with", []),
                    assigned_shifts=person_data.get("assigned_shifts", []),
                    fridays_count=person_data.get("fridays_count", 0),
                    saturdays_count=person_data.get("saturdays_count", 0),
                    sundays_count=person_data.get("sundays_count", 0),
                )
                self.people.append(person)

    def organize_schedule(self):
        # Create a dictionary with date as key and list of names as value
        date_schedule: Dict[str, List[str]] = {}
        for person in self.people:
            for shift in person.assigned_shifts:
                if shift not in date_schedule:
                    date_schedule[shift] = []
                date_schedule[shift].append(person.name)

        # Convert dates to datetime objects and sort them
        sorted_dates = sorted(
            [datetime.strptime(date, "%Y-%m-%d") for date in date_schedule.keys()]
        )

        # Organize into weeks
        for date in sorted_dates:
            week_num = date.isocalendar()[1]
            week_key = f"Week {week_num}"
            if week_key not in self.schedule:
                self.schedule[week_key] = {}
            day_name = date.strftime("%A")
            day_str = f"{day_name} ({date.strftime('%m-%d')})"
            assigned_names = date_schedule[date.strftime("%Y-%m-%d")]
            self.schedule[week_key][day_str] = assigned_names

    def create_schedule_sheet(self, wb: Workbook):
        ws = wb.create_sheet(title="Schedule")

        # Define styles
        week_header_font = Font(bold=True, size=14)
        day_header_font = Font(bold=True)
        assigned_fill = PatternFill(
            start_color="90EE90", end_color="90EE90", fill_type="solid"
        )  # Light Green
        weekend_fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )  # Gold
        assigned_weekend_fill = PatternFill(
            start_color="FFA500", end_color="FFA500", fill_type="solid"
        )  # Orange
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(
            horizontal="center", vertical="center"
        )  # Added Alignment

        current_row = 1

        # Sort weeks by week number
        sorted_weeks = sorted(self.schedule.keys(), key=lambda x: int(x.split()[1]))

        for week in sorted_weeks:
            # Write Week Header
            week_cell = ws.cell(row=current_row, column=1, value=week)
            week_cell.font = week_header_font

            # Determine the number of day columns to merge across
            day_headers = list(self.schedule[week].keys())
            num_days = len(day_headers)

            # Merge cells across all day columns for the week header
            if num_days > 1:
                ws.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=num_days,
                )
                week_cell = ws.cell(row=current_row, column=1, value=week)
                week_cell.font = week_header_font
                week_cell.alignment = center_alignment
            else:
                # If there's only one day, no need to merge, but still center the text
                week_cell.alignment = center_alignment

            current_row += 1

            # Write Day Headers
            for idx, day in enumerate(day_headers, start=1):
                day_cell = ws.cell(row=current_row, column=idx, value=day)
                day_cell.font = day_header_font
                day_cell.alignment = center_alignment

            current_row += 1

            # Prepare separate lists for first and second nurse names
            assigned_nurses_first = []
            assigned_nurses_second = []
            for day in day_headers:
                names = self.schedule[week][day]
                first_name = names[0] if len(names) >= 1 else ""
                second_name = names[1] if len(names) >= 2 else ""
                assigned_nurses_first.append(first_name)
                assigned_nurses_second.append(second_name)

            # Write First Nurse Names
            for idx, name in enumerate(assigned_nurses_first, start=1):
                cell = ws.cell(row=current_row, column=idx, value=name)
                # Extract day name and normalize
                day_name = day_headers[idx - 1].split()[0].title()
                # Determine fill based on day and assignment
                if day_name in {"Friday", "Saturday", "Sunday"}:
                    if name:
                        cell.fill = assigned_weekend_fill
                    else:
                        cell.fill = weekend_fill
                else:
                    if name:
                        cell.fill = assigned_fill
                # Apply border and alignment
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )  # Optional: Center text

            current_row += 1

            # Write Second Nurse Names
            for idx, name in enumerate(assigned_nurses_second, start=1):
                cell = ws.cell(row=current_row, column=idx, value=name)
                # Extract day name and normalize
                day_name = day_headers[idx - 1].split()[0].title()
                # Determine fill based on day and assignment
                if day_name in {"Friday", "Saturday", "Sunday"}:
                    if name:
                        cell.fill = assigned_weekend_fill
                    else:
                        cell.fill = weekend_fill
                else:
                    if name:
                        cell.fill = assigned_fill
                # Apply border and alignment
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )  # Optional: Center text

            current_row += 1

            # Add a blank row after each week for readability
            current_row += 1

        # Adjust column widths for better visibility
        for i, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_statistics_sheet(self, wb: Workbook):
        ws = wb.create_sheet(title="Statistics")

        # Define styles
        stats_header_font = Font(bold=True, size=12)
        day_header_font = Font(bold=True)
        stats_fill = PatternFill(
            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
        )  # Light Blue
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        current_row = 1

        # Write Stats Header
        ws.cell(row=current_row, column=1, value="Statistics")
        ws.cell(row=current_row, column=1).font = stats_header_font
        current_row += 1

        # Write Stats Table Headers
        headers = [
            "Name",
            "Fridays Assigned",
            "Saturdays Assigned",
            "Sundays Assigned",
            "Weekday Shifts",
            "Weekend Shifts",
            "Total Shifts",
        ]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=idx, value=header)
            cell.font = day_header_font
            cell.fill = stats_fill
            cell.border = thin_border
        current_row += 1

        # Populate Stats Data
        for person in self.people:
            ws.cell(row=current_row, column=1, value=person.name).border = thin_border
            ws.cell(row=current_row, column=2, value=person.fridays_count).border = (
                thin_border
            )
            ws.cell(row=current_row, column=3, value=person.saturdays_count).border = (
                thin_border
            )
            ws.cell(row=current_row, column=4, value=person.sundays_count).border = (
                thin_border
            )
            ws.cell(row=current_row, column=5, value=person.weekday_shifts).border = (
                thin_border
            )
            ws.cell(row=current_row, column=6, value=person.weekend_shifts).border = (
                thin_border
            )
            ws.cell(row=current_row, column=7, value=person.total_shifts).border = (
                thin_border
            )
            current_row += 1

        # Apply Borders to the stats table
        for row in ws.iter_rows(
            min_row=2, max_row=current_row - 1, min_col=1, max_col=7
        ):
            for cell in row:
                cell.border = thin_border

        # Adjust column widths for better visibility
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            for cell in column_cells:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[column].width = adjusted_width

    def create_spreadsheet(self):
        wb = Workbook()

        # Remove the default "Sheet"
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)

        # Create Schedule Sheet
        self.create_schedule_sheet(wb)

        # Create Statistics Sheet
        self.create_statistics_sheet(wb)
        wb.save(self.output_excel)
        logging.info(f"Schedule spreadsheet exported to:\n{self.output_excel}")

    def export(self):
        self.load_data()
        self.organize_schedule()
        self.create_spreadsheet()
