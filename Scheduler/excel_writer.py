from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys
import pandas as pd


def save_schedule_to_excel(
    assignments,
    weekend_counts,
    days_order,
    personas,
    filename="people_schedule.xlsx",
):
    df = pd.DataFrame(assignments)
    df["date_obj"] = pd.to_datetime(df["date"])
    df["week_number"] = df["date_obj"].dt.isocalendar().week
    df["weekday"] = df["date_obj"].dt.day_name()
    df = df.sort_values(by=["week_number", "date_obj"])

    # Group assignments by week
    weeks_data = {}
    for _, row in df.iterrows():
        w = int(row["week_number"])
        day = row["weekday"]  # e.g. "Monday", "Tuesday", ...
        date_str = row["date"]
        assigned_text = f"{row['person1']}\n{row['person2']}"

        if w not in weeks_data:
            weeks_data[w] = {}
        weeks_data[w][day] = {"date": date_str, "assignment": assigned_text}

    # Ensure each day is present with a default entry
    for week in weeks_data:
        for day in days_order:
            if day not in weeks_data[week]:
                weeks_data[week][day] = {
                    "date": "No Date",
                    "assignment": "No Assignment",
                }

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Schedule"
    row_idx = 1

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    holiday_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    unassigned_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    unassigned_both_fill = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )
    weekend_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )  # Light gray

    # Write schedule data week by week
    for week in sorted(weeks_data.keys()):
        # Merge cells for "Week X"
        sheet.merge_cells(
            start_row=row_idx,
            end_row=row_idx,
            start_column=1,
            end_column=len(days_order),
        )
        week_cell = sheet.cell(row=row_idx, column=1, value=f"Week {week}")
        week_cell.font = Font(bold=True)
        week_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        week_cell.border = thin_border
        row_idx += 1

        # Date row
        for col_idx, day_name in enumerate(days_order, start=1):
            date = weeks_data[week][day_name]["date"]
            date_cell = sheet.cell(row=row_idx, column=col_idx, value=date)
            date_cell.font = Font(bold=True)
            date_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            date_cell.border = thin_border
        row_idx += 1

        # Weekday header
        for col_idx, day_name in enumerate(days_order, start=1):
            header_cell = sheet.cell(row=row_idx, column=col_idx, value=day_name)
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            header_cell.border = thin_border
        row_idx += 1

        # Assigned pairs
        for col_idx, day_name in enumerate(days_order, start=1):
            data_dict = weeks_data[week][day_name]
            assignment = data_dict["assignment"]
            assignment_cell = sheet.cell(row=row_idx, column=col_idx, value=assignment)
            assignment_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            assignment_cell.border = thin_border

            # If it's one of Fri, Sat, Sun, fill background with a light gray color
            if day_name in ["Friday", "Saturday", "Sunday"]:
                assignment_cell.fill = weekend_fill

            # Override if holiday or unassigned
            if assignment == "Holiday":
                assignment_cell.fill = holiday_fill
            elif "\nUnassigned" in assignment:
                if assignment.count("Unassigned") == 2:
                    assignment_cell.fill = unassigned_both_fill
                else:
                    assignment_cell.fill = unassigned_fill

        row_idx += 1

    # Stats Section
    row_idx += 1
    stats_title = f"Weekend Assignment Stats (Total Weeks: {len(weeks_data)})"
    sheet.merge_cells(start_row=row_idx, end_row=row_idx, start_column=1, end_column=8)
    stats_cell = sheet.cell(row=row_idx, column=1, value=stats_title)
    stats_cell.font = Font(bold=True)
    stats_cell.alignment = Alignment(horizontal="left", vertical="center")
    row_idx += 1

    headers = [
        "Person Name",
        "Fridays Assigned",
        "Saturdays Assigned",
        "Sundays Assigned",
        "Total Weekend Assignments",
        "Week Numbers",
        "Unavailable Dates",
        "Incompatible With",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
    row_idx += 1

    # Fill in stats
    for person in personas:
        name = person["name"]
        unavail_dates = ", ".join(person.get("unavailable_days", []))
        incompatible_with = ", ".join(person.get("incompatible_with", []))
        wc = weekend_counts.get(name, {})
        fridays_assigned = wc.get("fridays", 0)
        saturdays_assigned = wc.get("saturdays", 0)
        sundays_assigned = wc.get("sundays", 0)
        total_weekend = wc.get("count", 0)
        week_numbers = ", ".join(map(str, sorted(set(wc.get("weeks", [])))))

        stats_row = [
            name,
            fridays_assigned,
            saturdays_assigned,
            sundays_assigned,
            total_weekend,
            week_numbers,
            unavail_dates,
            incompatible_with,
        ]
        for col_idx, val in enumerate(stats_row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
        row_idx += 1

    # Adjust columns
    for i, day_name in enumerate(days_order, start=1):
        letter = get_column_letter(i)
        sheet.column_dimensions[letter].width = 20

    stats_widths = [20, 20, 20, 20, 25, 30, 30, 30]
    for i, w in enumerate(stats_widths, start=1):
        letter = get_column_letter(i)
        sheet.column_dimensions[letter].width = w

    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 30

    try:
        wb.save(filename)
        print(f"Schedule successfully generated and saved as '{filename}'.")
    except Exception as e:
        print(f"Failed to save the schedule: {e}")
        sys.exit(1)
