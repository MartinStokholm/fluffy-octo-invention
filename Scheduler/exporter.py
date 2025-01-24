import json
import os
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import logging


from utils import ensure_dir_exists, get_relative_path
from typing import Dict, List
from person import Person
from pathlib import Path
from openpyxl import Workbook
from datetime import datetime
from collections import defaultdict
from dataclasses import dataclass, asdict
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


@dataclass
class HolidaysResult:
    holiday_name: str
    people_names: List[str]
    date: List[str]


@dataclass
class PersonResult:
    name: str
    working_day: str
    absence_days: List[str]
    incompatible_with: List[str]
    assigned_shifts: List[str]
    fridays_count: int
    saturdays_count: int
    sundays_count: int

    @property
    def weekday_shifts(self) -> int:
        """
        Calculate the number of assigned shifts excluding weekends.
        """
        return len(self.assigned_shifts) - (
            self.fridays_count + self.saturdays_count + self.sundays_count
        )

    @property
    def weekend_shifts(self) -> int:
        """
        Calculate the number of assigned shifts on weekends.
        """
        return self.fridays_count + self.saturdays_count + self.sundays_count

    @property
    def total_shifts(self) -> int:
        """
        Calculate the total number of assigned shifts.
        """
        return len(self.assigned_shifts)


class SpreadsheetExporter:
    def __init__(
        self,
        json_filepath: str,
        output_excel: str,
        base_dir: Path,
        holidays: List[HolidaysResult] = [],
    ):
        self.json_filepath = Path(json_filepath)
        self.output_excel = Path(output_excel)
        self.base_dir = base_dir
        self.people: List[PersonResult] = []
        self.schedule: Dict[str, Dict[str, List[str]]] = {}  # {week: {day: [names]}}
        self.holidays: List[datetime] = holidays

    def load_data(self):
        with self.json_filepath.open("r") as file:
            data = json.load(file)
            for person_data in data:
                person = PersonResult(
                    name=person_data["name"],
                    working_day=person_data["working_day"],
                    absence_days=person_data.get("absence_days", []),
                    incompatible_with=person_data.get("incompatible_with", []),
                    assigned_shifts=person_data.get("assigned_shifts", []),
                    fridays_count=person_data.get("fridays_count", 0),
                    saturdays_count=person_data.get("saturdays_count", 0),
                    sundays_count=person_data.get("sundays_count", 0),
                )
                self.people.append(person)

    def organize_schedule(self):
        # Create a dictionary with date as key and list of names as value
        date_schedule: Dict[str, List[str]] = {}
        for person in self.people:
            for shift in person.assigned_shifts:
                if shift not in date_schedule:
                    date_schedule[shift] = []
                date_schedule[shift].append(person.name)

        # Convert dates to datetime objects and sort them
        sorted_dates = sorted(
            [datetime.strptime(date, "%Y-%m-%d") for date in date_schedule.keys()]
        )

        # Organize into weeks
        for date in sorted_dates:
            week_num = date.isocalendar()[1]
            week_key = f"Week {week_num}"
            if week_key not in self.schedule:
                self.schedule[week_key] = {}
            day_name = date.strftime("%A")
            day_str = f"{day_name} ({date.strftime('%m-%d')})"
            assigned_names = date_schedule[date.strftime("%Y-%m-%d")]
            self.schedule[week_key][day_str] = assigned_names

        # Sort days within each week
        for week in self.schedule:
            sorted_days = sorted(
                self.schedule[week].keys(),
                key=lambda x: datetime.strptime(x.split()[1].strip("()"), "%m-%d"),
            )
            self.schedule[week] = {day: self.schedule[week][day] for day in sorted_days}

    def create_schedule_sheet(self, wb: Workbook):
        ws = wb.create_sheet(title="Schedule")

        # Define styles
        week_header_font = Font(bold=True, size=14)
        day_header_font = Font(bold=True)

        assigned_fill = PatternFill(
            start_color="FCFCBF", end_color="FCFCBF", fill_type="solid"
        )  # Weekdays

        holiday_fill = PatternFill(
            start_color="FCBFFC", end_color="FCBFFC", fill_type="solid"
        )  # Holidays

        weekend_fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )  # Gold

        assigned_friday_fill = PatternFill(
            start_color="87CEEB", end_color="87CEEB", fill_type="solid"
        )  # Fridays
        assigned_saturday_fill = PatternFill(
            start_color="FA8072", end_color="FA8072", fill_type="solid"
        )  # Saturdays
        assigned_sunday_fill = PatternFill(
            start_color="90EE90", end_color="90EE90", fill_type="solid"
        )  # Sundays

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(
            horizontal="center", vertical="center"
        )  # Added Alignment

        current_row = 1

        # Sort weeks by week number
        sorted_weeks = sorted(self.schedule.keys(), key=lambda x: int(x.split()[1]))

        for week in sorted_weeks:
            # Write Week Header
            week_cell = ws.cell(row=current_row, column=1, value=week)
            week_cell.font = week_header_font

            # Determine the number of day columns to merge across
            day_headers = list(self.schedule[week].keys())
            num_days = len(day_headers)

            # Merge cells across all day columns for the week header
            if num_days > 1:
                ws.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=num_days,
                )
                week_cell = ws.cell(row=current_row, column=1, value=week)
                week_cell.font = week_header_font
                week_cell.alignment = center_alignment
            else:
                # If there's only one day, no need to merge, but still center the text
                week_cell.alignment = center_alignment

            current_row += 1

            # Write Day Headers
            for idx, day in enumerate(day_headers, start=1):
                day_cell = ws.cell(row=current_row, column=idx, value=day)
                day_cell.font = day_header_font
                day_cell.alignment = center_alignment

            current_row += 1

            # Prepare separate lists for first and second nurse names
            assigned_nurses_first = []
            assigned_nurses_second = []
            for day in day_headers:
                names = self.schedule[week][day]
                first_name = names[0] if len(names) >= 1 else ""
                second_name = names[1] if len(names) >= 2 else ""
                assigned_nurses_first.append(first_name)
                assigned_nurses_second.append(second_name)

            # Write First Nurse Names
            for idx, name in enumerate(assigned_nurses_first, start=1):
                cell = ws.cell(row=current_row, column=idx, value=name)
                # Extract day name and normalize
                day_name = day_headers[idx - 1].split()[0].title()
                # Determine fill based on day and assignment
                day_date_str = day_headers[idx - 1].split()[1].strip("()")
                day_date = datetime.strptime(day_date_str, "%m-%d")

                # Apply holiday fill if the day is a holiday
                if any(
                    datetime.strptime(holiday["date"], "%Y-%m-%d").strftime("%m-%d")
                    == day_date.strftime("%m-%d")
                    for holiday in self.holidays
                ):
                    cell.fill = holiday_fill
                else:
                    if day_name == "Friday":
                        if name:
                            cell.fill = assigned_friday_fill
                        else:
                            cell.fill = weekend_fill
                    elif day_name == "Saturday":
                        if name:
                            cell.fill = assigned_saturday_fill
                        else:
                            cell.fill = weekend_fill
                    elif day_name == "Sunday":
                        if name:
                            cell.fill = assigned_sunday_fill
                        else:
                            cell.fill = weekend_fill
                    else:
                        if name:
                            cell.fill = assigned_fill
                # Apply border and alignment
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )  # Optional: Center text

            current_row += 1

            # Write Second Nurse Names
            for idx, name in enumerate(assigned_nurses_second, start=1):
                cell = ws.cell(row=current_row, column=idx, value=name)
                # Extract day name and normalize
                day_name = day_headers[idx - 1].split()[0].title()
                # Determine fill based on day and assignment
                day_date_str = day_headers[idx - 1].split()[1].strip("()")
                day_date = datetime.strptime(day_date_str, "%m-%d")

                # Apply holiday fill if the day is a holiday
                if any(
                    datetime.strptime(holiday["date"], "%Y-%m-%d").strftime("%m-%d")
                    == day_date.strftime("%m-%d")
                    for holiday in self.holidays
                ):
                    cell.fill = holiday_fill
                else:
                    if day_name == "Friday":
                        if name:
                            cell.fill = assigned_friday_fill
                        else:
                            cell.fill = weekend_fill
                    elif day_name == "Saturday":
                        if name:
                            cell.fill = assigned_saturday_fill
                        else:
                            cell.fill = weekend_fill
                    elif day_name == "Sunday":
                        if name:
                            cell.fill = assigned_sunday_fill
                        else:
                            cell.fill = weekend_fill
                    else:
                        if name:
                            cell.fill = assigned_fill
                # Apply border and alignment
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )  # Optional: Center text

            current_row += 1

            # Add a blank row after each week for readability
            current_row += 1

        # Adjust column widths for better visibility
        for i, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_statistics_sheet(self, wb: Workbook):
        ws = wb.create_sheet(title="Statistics")

        # Define styles
        stats_header_font = Font(bold=True, size=12)
        day_header_font = Font(bold=True)
        stats_fill = PatternFill(
            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
        )  # Light Blue
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        current_row = 1

        # Write Stats Header
        ws.cell(row=current_row, column=1, value="")
        ws.cell(row=current_row, column=1).font = stats_header_font
        current_row += 1

        # Write Stats Table Headers
        headers = [
            "Name",
            "Fridays Assigned",
            "Saturdays Assigned",
            "Sundays Assigned",
            "Weekday Shifts",
            "Weekend Shifts",
            "Total Shifts",
        ]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=idx, value=header)
            cell.font = day_header_font
            cell.fill = stats_fill
            cell.border = thin_border
        current_row += 1

        # Populate Stats Data
        for person in self.people:
            ws.cell(row=current_row, column=1, value=person.name).border = thin_border
            ws.cell(row=current_row, column=2, value=person.fridays_count).border = (
                thin_border
            )
            ws.cell(row=current_row, column=3, value=person.saturdays_count).border = (
                thin_border
            )
            ws.cell(row=current_row, column=4, value=person.sundays_count).border = (
                thin_border
            )
            ws.cell(row=current_row, column=5, value=person.weekday_shifts).border = (
                thin_border
            )
            ws.cell(row=current_row, column=6, value=person.weekend_shifts).border = (
                thin_border
            )
            ws.cell(row=current_row, column=7, value=person.total_shifts).border = (
                thin_border
            )
            current_row += 1

        # Apply Borders to the stats table
        for row in ws.iter_rows(
            min_row=2, max_row=current_row - 1, min_col=1, max_col=7
        ):
            for cell in row:
                cell.border = thin_border

        # Adjust column widths for better visibility
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            for cell in column_cells:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[column].width = adjusted_width

    def create_spreadsheet(self):
        wb = Workbook()

        # Remove the default "Sheet"
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)

        # Create Schedule Sheet
        self.create_schedule_sheet(wb)

        # Create Statistics Sheet
        self.create_statistics_sheet(wb)
        try:
            wb.save(self.output_excel)
            relative_excel_path = get_relative_path(self.output_excel, self.base_dir)
            logging.info(f"💾 Output: {relative_excel_path}")
        except Exception as e:
            relative_excel_path = get_relative_path(self.output_excel, self.base_dir)
            logging.error(
                f"⚠️ Failed to export spreadsheet to {relative_excel_path}: {e}"
            )

    def export(self):
        self.load_data()
        self.organize_schedule()
        self.create_spreadsheet()


class GraphExporter:
    def __init__(self, json_filepath: str, output_graph: str, base_dir: Path):
        self.json_filepath = Path(json_filepath)
        self.base_dir = base_dir
        self.output_graph = self.base_dir / output_graph
        self.people: List[Dict] = []
        self.assignments: Dict[str, Dict[str, int]] = defaultdict(
            lambda: {"Friday": 0, "Saturday": 0, "Sunday": 0}
        )
        self.graph_path = self.output_graph

    def load_data(self):
        with self.json_filepath.open("r") as file:
            data = json.load(file)
            for person_data in data:
                person = {
                    "name": person_data["name"],
                    "fridays_count": person_data.get("fridays_count", 0),
                    "saturdays_count": person_data.get("saturdays_count", 0),
                    "sundays_count": person_data.get("sundays_count", 0),
                }
                self.people.append(person)

    def organize_data(self):
        for person in self.people:
            self.assignments[person["name"]]["Friday"] = person["fridays_count"]
            self.assignments[person["name"]]["Saturday"] = person["saturdays_count"]
            self.assignments[person["name"]]["Sunday"] = person["sundays_count"]

    def plot_distribution_comparison(self):
        # Prepare data for plotting
        persons = list(self.assignments.keys())
        fridays = [self.assignments[person]["Friday"] for person in persons]
        saturdays = [self.assignments[person]["Saturday"] for person in persons]
        sundays = [self.assignments[person]["Sunday"] for person in persons]

        x = range(len(persons))  # the label locations
        width = 0.25  # the width of the bars

        plt.figure(figsize=(20, 10))
        plt.bar(
            [i - width for i in x],
            fridays,
            width,
            label="Fridays",
            color="#87CEEB",  # skyblue
        )
        plt.bar(x, saturdays, width, label="Saturdays", color="#FA8072")  # salmon
        plt.bar(
            [i + width for i in x],
            sundays,
            width,
            label="Sundays",
            color="#90EE90",  # lightgreen
        )

        plt.xlabel("Person", fontsize=14)
        plt.ylabel("Number of Assignments", fontsize=14)
        plt.title("Weekend Shift Assignments per Person", fontsize=16)
        plt.xticks(x, persons, rotation=45, ha="right")
        plt.legend()
        plt.tight_layout()

        plt.savefig(self.graph_path)
        plt.close()
        relative_graph_path = get_relative_path(self.graph_path, self.base_dir)
        logging.info(f"💾 Output: {relative_graph_path}")

    def export(self):
        self.load_data()
        self.organize_data()
        self.plot_distribution_comparison()


class JsonExporter:
    def __init__(
        self,
        output_filepath: str = "data/people_assigned.json",
        base_dir: Path = Path("."),
    ):
        self.output_filepath = Path(output_filepath)
        self.base_dir = base_dir

    def export(self, people: List[Person]):
        """
        Saves the assigned shifts and weekend counts for each person to a JSON file.

        :param people: List of Person instances with assigned shifts.
        """
        assignments = []
        for person in people:
            assigned_dates = [
                shift_date.strftime("%Y-%m-%d") for shift_date in person.schedule
            ]
            person_data = {
                "name": person.name,
                "working_day": person.working_day,
                "absence_days": person.absence_days,
                "incompatible_with": person.incompatible_with,
                "assigned_shifts": assigned_dates,
                "fridays_count": person.fridays_count,
                "saturdays_count": person.saturdays_count,
                "sundays_count": person.sundays_count,
            }
            assignments.append(person_data)
        try:
            with self.output_filepath.open("w", encoding="utf-8") as outfile:
                json.dump(assignments, outfile, indent=4, ensure_ascii=False)
            logging.info(
                f"💾 Output: {get_relative_path(self.output_filepath, self.base_dir)}."
            )
        except Exception as e:
            logging.error(
                f"❌ Output: {get_relative_path(self.output_filepath, self.base_dir)}: {e}"
            )
